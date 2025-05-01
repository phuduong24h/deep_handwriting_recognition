import cv2
from PIL import Image
import pytesseract
import os
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import scrolledtext

# Danh sách thuốc
training_texts = [
    'Paracetamol', 'Aspirin', 'Ibuprofen', 'Amoxicillin', 'Cefuroxime', 'Vitamin C', 'Panadol', 'Decolgen', 'Tiffy', 'Salonpas',
    'Berberin', 'Smecta', 'Efferalgan', 'Hapacol', 'Coldacmin', 'Omeprazole', 'Loperamide', 'Betadine', 'Hydroxyzine', 'Loratadine',
    'Zyrtec', 'Alaxan', 'Magne B6', 'Spasfon', 'Antot', 'Rhumenol', 'Nospa', 'Tetracycline', 'Clarithromycin', 'Doxycycline',
    'Enterogermina', 'Becozyme', 'Neocodion', 'Tylenol', 'Morphine', 'Codeine', 'Cotrimoxazole', 'Ampicillin', 'Metronidazole',
    'Clorpheniramine', 'Acetaminophen', 'Rifampicin', 'Isoniazid', 'Streptomycin', 'Azithromycin', 'Ciprofloxacin', 'Ketorolac', 'Naproxen', 'Rabeprazole', 'Thuốc ho'
]

# Gán nhãn cho thuốc
training_labels = ['Thuốc'] * len(training_texts)

# Khởi tạo vectorizer và mô hình KNN
vectorizer = TfidfVectorizer()
X_train = vectorizer.fit_transform(training_texts)
knn_image = KNeighborsClassifier(n_neighbors=3, weights='distance')
knn_image.fit(X_train, training_labels)

# Đường dẫn Tesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

# Hàm nhận diện văn bản từ ảnh
def tesseract(image_path):
    img = Image.open(image_path)
    config = '--psm 6 --oem 3 lang=eng'
    text = pytesseract.image_to_string(img, config=config)
    return text.strip()

# Hàm phân loại văn bản
def classify_text(text):
    X_test = vectorizer.transform([text])
    similarity_scores = cosine_similarity(X_test, X_train)
    max_score = np.max(similarity_scores)

    if max_score >= 0.8:
        predicted_class = knn_image.predict(X_test)[0]
    else:
        predicted_class = 'Không phải thuốc'

    return predicted_class

# Lưu kết quả vào Excel chỉ khi là "Thuốc"
def save_to_excel(text, category, excel_path='thuoc_detected.xlsx'):
    if category == 'Thuốc':  # Chỉ lưu khi phân loại là thuốc
        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['ID', 'Tên nhận diện', 'Phân loại'])

        new_row = [sheet.max_row, text, category]
        sheet.append(new_row)
        workbook.save(excel_path)
        return f"Đã lưu kết quả vào file: {excel_path}"
    else:
        return "Không lưu vì không phải thuốc"

# Hàm để xử lý ảnh và nhận diện văn bản
def process_image_for_text(image_path, gui_output):
    recognized_text = tesseract(image_path)
    lines = recognized_text.split('\n')  # Tách các dòng văn bản
    output = ""
    for line in lines:
        category = classify_text(line)  # Phân loại từng dòng
        output += f"Dòng: {line} | Phân loại: {category}\n"
        
        save_message = save_to_excel(line, category)  # Lưu kết quả vào Excel nếu là thuốc
        output += save_message + "\n"
    
    # Hiển thị kết quả lên GUI
    gui_output.insert(tk.END, output)
    gui_output.yview(tk.END)  

# Hàm để hiển thị webcam và lưu ảnh
def show_webcam(gui_output):
    camera = cv2.VideoCapture(0)
    if not camera.isOpened():
        print("Không thể mở webcam. Kiểm tra lại thiết bị.")
        exit()

    while True:
        ret, frame = camera.read()
        if not ret:
            print("Không thể lấy hình ảnh từ webcam.")
            break

        cv2.imshow('Text detection', frame)
        key = cv2.waitKey(1) & 0xFF
        if key == ord('s'):  # Nhấn 's' để lưu ảnh
            image_path = 'test1.jpg'
            cv2.imwrite(image_path, frame)
            print(f"Ảnh đã được lưu thành {image_path}")
            process_image_for_text(image_path, gui_output)  # Xử lý ảnh và hiển thị kết quả phân loại lên GUI
            camera.release()  # Tắt camera sau khi nhận diện
            cv2.destroyAllWindows()  # Đóng cửa sổ webcam
            break  # Thoát khỏi vòng lặp sau khi nhấn 's'

        cv2.imshow('Webcam Feed', frame)  # Hiển thị ảnh webcam liên tục

# Hàm tạo GUI
def create_gui():
    # Tạo cửa sổ Tkinter
    window = tk.Tk()
    window.title("Phân loại thuốc từ ảnh")

    # Tạo Text Widget để hiển thị kết quả
    gui_output = scrolledtext.ScrolledText(window, width=80, height=20, wrap=tk.WORD)
    gui_output.pack(padx=10, pady=10)

    # Nút để bắt đầu hiển thị webcam và lưu ảnh
    start_button = tk.Button(window, text="Bắt đầu nhận diện", command=lambda: show_webcam(gui_output))
    start_button.pack(pady=10)

    # Chạy GUI
    window.mainloop()

# Bắt đầu GUI
create_gui()
