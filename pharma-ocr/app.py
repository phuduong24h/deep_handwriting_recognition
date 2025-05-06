import cv2
from PIL import Image
import pytesseract
import os
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook, load_workbook
import pandas as pd
import tkinter as tk
from tkinter import scrolledtext

# Đường dẫn đến Tesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"

# Load dữ liệu huấn luyện từ danh sách thuốc
thuoc_df = pd.read_excel("danh_sach_thuoc_nhieu_trang.xlsx")
if "Tên thuốc" not in thuoc_df.columns or "Hoạt chất" not in thuoc_df.columns:
    raise ValueError("File 'danh_sach_thuoc_nhieu_trang.xlsx' phải có cột 'Tên thuốc' và 'Hoạt chất'.")

training_texts = [str(text).strip().lower() for text in thuoc_df["Tên thuốc"].dropna()]
training_labels = ['Thuốc'] * len(training_texts)

# Load dữ liệu công dụng
hoatchat_df = pd.read_excel("100_hoat_chat_va_nhom_benh.xlsx")
hoatchat_df.fillna("", inplace=True)

# Vector hóa dữ liệu
vectorizer = TfidfVectorizer()
X_train = vectorizer.fit_transform(training_texts)

# Huấn luyện mô hình KNN
knn_image = KNeighborsClassifier(n_neighbors=3, weights='distance')
knn_image.fit(X_train, training_labels)

# OCR ảnh
def tesseract(image_path):
    img = Image.open(image_path)
    config = '--psm 6 --oem 3 -l eng'
    text = pytesseract.image_to_string(img, config=config)
    return text.strip().lower()

# Phân loại văn bản
def classify_text(text):
    text = text.lower()
    X_test = vectorizer.transform([text])
    similarity_scores = cosine_similarity(X_test, X_train)
    max_score = np.max(similarity_scores)
    best_match_index = np.argmax(similarity_scores)
    best_match = training_texts[best_match_index]

    if max_score >= 0.8:
        predicted_class = 'Thuốc'
    elif max_score >= 0.5:
        predicted_class = 'Gần giống'
    else:
        predicted_class = 'Không phải thuốc'

    return predicted_class, max_score, best_match

# Truy xuất hoạt chất
def get_hoat_chat(ten_thuoc):
    matches = thuoc_df[thuoc_df["Tên thuốc"].str.lower().str.strip() == ten_thuoc.strip().lower()]
    if not matches.empty:
        return matches.iloc[0]["Hoạt chất"]
    return ""

# Truy xuất công dụng
def get_cong_dung(hoat_chat):
    if not hoat_chat:
        return ""
    matches = hoatchat_df[hoatchat_df["Hoạt chất"].str.lower().str.strip() == hoat_chat.strip().lower()]
    if not matches.empty:
        return matches.iloc[0]["Điều trị"]
    return ""

# Lưu kết quả nếu là thuốc
def save_to_excel(text, category, excel_path='thuoc_detected.xlsx'):
    if category == 'Thuốc':
        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['ID', 'Tên nhận diện', 'Phân loại'])

        new_row = [sheet.max_row, text, category]
        sheet.append(new_row)
        workbook.save(excel_path)
        return f"✅ Đã lưu vào file: {excel_path}"
    else:
        return "❌ Không lưu vì không phải thuốc"

# Xử lý ảnh và in kết quả vào GUI
def process_image_for_text(image_path, gui_output):
    recognized_text = tesseract(image_path)
    lines = recognized_text.split('\n')
    found = False  # Kiểm tra xem có dòng nào hợp lệ hay không

    for line in lines:
        if not line.strip():
            continue
        category, score, best_match = classify_text(line)
        
        if score >= 0.5:
            found = True
            hoat_chat = get_hoat_chat(best_match)
            cong_dung = get_cong_dung(hoat_chat)

            result_msg = f"Dòng: {line} | Phân loại: {category} | Gần giống: {best_match} ({score:.2f})"
            if hoat_chat:
                result_msg += f" | Hoạt chất: {hoat_chat}"
            if cong_dung:
                result_msg += f" | Điều trị: {cong_dung}"

            if category == 'Thuốc':
                gui_output.insert(tk.END, result_msg + "\n", 'green')
            elif category == 'Gần giống':
                gui_output.insert(tk.END, result_msg + "\n", 'red')
            else:
                gui_output.insert(tk.END, result_msg + "\n", 'black')

            save_message = save_to_excel(line, category)
            gui_output.insert(tk.END, save_message + "\n")
            gui_output.yview(tk.END)

    if not found:
        gui_output.insert(tk.END, "⚠️ Không có tên thuốc nào trong ảnh.\n", 'red')
        gui_output.yview(tk.END)

# Mở webcam và nhận diện ảnh
def show_webcam(gui_output):
    camera = cv2.VideoCapture(0)
    if not camera.isOpened():
        print("❌ Không thể mở webcam.")
        return

    while True:
        ret, frame = camera.read()
        if not ret:
            print("❌ Không thể lấy hình ảnh từ webcam.")
            break

        cv2.imshow('Webcam - Nhấn S để nhận diện', frame)
        key = cv2.waitKey(1) & 0xFF
        if key == ord('s'):
            image_path = 'test1.jpg'
            cv2.imwrite(image_path, frame)
            print(f"📸 Ảnh đã lưu: {image_path}")
            process_image_for_text(image_path, gui_output)
            break

    camera.release()
    cv2.destroyAllWindows()

# GUI chính
def create_gui():
    window = tk.Tk()
    window.title("🧪 Nhận diện thuốc từ ảnh")

    gui_output = scrolledtext.ScrolledText(window, width=100, height=25, wrap=tk.WORD)
    gui_output.pack(padx=10, pady=10)
    gui_output.tag_configure('green', foreground='green')
    gui_output.tag_configure('red', foreground='red')
    gui_output.tag_configure('black', foreground='black')

    start_button = tk.Button(window, text="Bắt đầu nhận diện (mở webcam)", command=lambda: show_webcam(gui_output))
    start_button.pack(pady=10)

    window.mainloop()

# Chạy chương trình
if __name__ == "__main__":
    create_gui()
